import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from PIL import Image, ImageTk
import random
import math
import json
import customtkinter as ctk
from RangeSlider.RangeSlider import RangeSliderH
from scipy.spatial import Delaunay
from openpyxl import Workbook

# Constants
SQUARE_SIZE = 1000
IMG_PATH = "C:\\Python\\Project_ACMN\\"
MAP_BG_FILENAME = "map_bg.png"
STATION_IMG_FILENAME = 'station.ico'
CITY_IMG_FILENAME = "city1.png"
SCALE_OPTIONS = {
    '100000m': 1000 / 100000,
    '200000m': 1000 / 200000,
    '50000m': 1000 / 50000
}

# Data
city_centers = []
base_stations = []

# Initialize Tkinter root
ctk.set_appearance_mode("Light")  # Set theme ("System", "Dark", "Light")
ctk.set_default_color_theme("blue")  # Set color theme

root = ctk.CTk()
root.iconbitmap(default=IMG_PATH + STATION_IMG_FILENAME)

root.title("Adjacent Cells in Mobile Networks")
root.state('zoomed')
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()

root.geometry(f"{screen_width}x{screen_height}")
canvas = tk.Canvas(root, bg='white', width=995, height=995)

canvas.grid(row=0, column=0, rowspan=22, sticky="n")

# Load and resize images
img = Image.open(IMG_PATH + MAP_BG_FILENAME).resize((980, 980))
img_tk = ImageTk.PhotoImage(img)
city_icon = Image.open(IMG_PATH + CITY_IMG_FILENAME).resize((30, 30))
city_icon_tk = ImageTk.PhotoImage(city_icon)

canvas.create_rectangle(5, 5, 995,995 , fill='white', width=5, tags="static")
canvas.create_image(10, 10, anchor=tk.NW, image=img_tk, tags="static")

def are_points_within_range(x, y, points, min_distance, max_distance=None):
    '''This function checks if a given point (x, y) is within a certain distance range of any points from the list.'''
    for point in points:
        px, py = point[:2]
        distance = math.hypot(x - px, y - py)  # More efficient calculation
        if max_distance is not None:
            if min_distance <= distance <= max_distance:
                return True
        elif distance < min_distance:
            return True
    return False

def on_canvas_click(event):
    for idx, station in enumerate(base_stations):
        x, y, radius = station["x"], station["y"], station["radius"]
        if (x - event.x)**2 + (y - event.y)**2 <= radius**2:
            highlight_station(idx)
            break
canvas.bind("<Button-1>", on_canvas_click)

def highlight_station(idx):
    # Highlight the station on the canvas
    station = base_stations[idx]
    x, y, radius = station["x"], station["y"], station["radius"]
    if current_zoom == 1.0:
        canvas.create_oval(x - radius, y - radius, x + radius, y + radius, outline='yellow', width=4, tags=("zoomable","highlight_{}".format(station["id"])))
    # Check and highlight in the main station list ('tree')
    for i in tree.get_children():
        if tree.item(i)['values'][0] == station["id"]:
            tree.selection_set(i)  # This highlights the row in 'tree'
            tree.see(i)  # This ensures the row is visible in 'tree'
            break

    for i in selected_tree.get_children():
        if selected_tree.item(i)['values'][0] == station["id"]:
            selected_tree.selection_set(i)
            selected_tree.see(i)
            selected_tree.move(i, selected_tree.parent(i), 0)
            return  # If found, no need to add it again
    # If station is not found in the selected stations list, add it
    new_item = selected_tree.insert("", 0, values=(station["id"], round(x), round(y), round(station["radius"] * METERS_PER_PIXEL), station["position"]))
    selected_tree.focus(new_item)
    selected_tree.selection_set(new_item)
    selected_tree.see(new_item)

def clear_highlight():
    '''Clear the highlighted station on the canvas and deselect any row in the table.'''
    all_items = canvas.find_all()
    for item in all_items:
        tags = canvas.gettags(item)
        for tag in tags:
            if 'highlight_' in tag:
                canvas.delete(item)
                break
    # Deselect in both tree views
    tree.selection_remove(tree.selection())
    selected_tree.selection_remove(selected_tree.selection())

    for item in selected_tree.get_children():
        selected_tree.delete(item)

def remove_highlight_by_id(station_id):
    # Remove the highlight from the canvas by the station's ID
    for idx, station in enumerate(base_stations):
        if station["id"] == station_id:
            canvas.delete("highlight_{}".format(station_id))
            break

def delete_selected_station():
    selected_item = selected_tree.selection()
    if selected_item:
        item_values = selected_tree.item(selected_item)["values"]   # Get the item's values which contains the station ID as the first element
        station_id = item_values[0]
        remove_highlight_by_id(station_id)  # Remove the highlight for this station
        selected_tree.delete(selected_item)  # Remove the item from the tree

def find_station():
    station_id = station_id_entry_var.get()
    if station_id.isdigit():
        station_id = int(station_id)
        for idx, station in enumerate(base_stations):
            if station["id"] == station_id:
                highlight_station(idx)
                for i in triangulation_table.get_children():
                    if triangulation_table.item(i)['values'][0] == station["id"]:
                        triangulation_table.selection_set(i)  # This highlights the row in 'tree'
                        triangulation_table.see(i)  # This ensures the row is visible in 'tree'
                        break
                for i in neighbors_table.get_children():
                    if neighbors_table.item(i)['values'][0] == station["id"]:
                        neighbors_table.selection_set(i)  # This highlights the row in 'tree'
                        neighbors_table.see(i)  # This ensures the row is visible in 'tree'
                        break
                break
        else:
            tk.messagebox.showinfo("Information", "Station ID not found.")
    else:
        tk.messagebox.showwarning("Warning", "Please enter a valid station ID.")

current_zoom = 1.0
MIN_ZOOM_LEVEL = 1.0  # Constants to define the min and max zoom levels
MAX_ZOOM_LEVEL = 7.0  # Adjust as needed
def zoom(event):
    global current_zoom
    scale_factor = 1.1
    if event.delta > 0:  # scroll up, zoom in
        new_zoom = current_zoom * scale_factor
        if new_zoom < MAX_ZOOM_LEVEL:
            canvas.scale("zoomable", event.x, event.y, scale_factor, scale_factor)
            current_zoom = new_zoom
    elif event.delta < 0:  # scroll down, zoom out
        new_zoom = current_zoom / scale_factor
        if new_zoom <= MIN_ZOOM_LEVEL:
            # If new zoom level is at or below minimum, reset to exactly 1.0
            reset_scale = 1.0 / current_zoom
            canvas.scale("zoomable", event.x, event.y, reset_scale, reset_scale)
            current_zoom = 1.0
        else:
            canvas.scale("zoomable", event.x, event.y, 1/scale_factor, 1/scale_factor)
            current_zoom = new_zoom

canvas.bind("<MouseWheel>", zoom)
last_drag_position = None

def start_pan(event):
    global last_drag_position
    last_drag_position = (event.x, event.y)
def do_pan(event):
    global last_drag_position
    dx = event.x - last_drag_position[0]
    dy = event.y - last_drag_position[1]
    canvas.move("zoomable", dx, dy)
    last_drag_position = (event.x, event.y)
def end_pan(event):
    global last_drag_position
    last_drag_position = None

canvas.bind("<ButtonPress-3>", start_pan)
canvas.bind("<B3-Motion>", do_pan)
canvas.bind("<ButtonRelease-3>", end_pan)

def save_configuration_as():
    filename = tk.filedialog.asksaveasfilename(
        defaultextension=".json",
        filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
        title="Save Configuration As..."
    )
    if filename:
        try:
            config = {
                'cities': city_centers,
                'stations': base_stations,
                'selected_scale': scale_selection_var.get()
            }
            with open(filename, 'w') as f:
                json.dump(config, f, indent=4)
            tk.messagebox.showinfo("Save Configuration", "Configuration saved successfully.")
        except Exception as e:
            tk.messagebox.showerror("Save Configuration", f"An error occurred while saving: {e}")

def load_configuration():
    filename = tk.filedialog.askopenfilename(
        filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
        title="Load Configuration"
    )
    if filename:
        try:
            with open(filename, 'r') as f:
                config = json.load(f)
                global city_centers, base_stations
                city_centers = config['cities']
                base_stations = config['stations']
                scale_selection_var.set(config.get('selected_scale', '77500m'))
                on_scale_select(None)  # Update the scale based on the loaded configuration
                draw_from_loaded_data()
        except FileNotFoundError:
            tk.messagebox.showwarning("Load Configuration", "No configuration file found.")
        except json.JSONDecodeError:
            tk.messagebox.showerror("Load Configuration", "Configuration file is corrupted.")
        except Exception as e:
            tk.messagebox.showerror("Load Configuration", f"An error occurred while loading: {e}")

def draw_from_loaded_data():
    # Clear the canvas first
    canvas.delete("city")
    canvas.delete("base_station")
    canvas.delete("triangulation")
    triangulation_table.delete(*triangulation_table.get_children())
    clear_neighbors()
    clear_highlight()

    for i in tree.get_children():
        tree.delete(i)
    for i in city_tree.get_children():
        city_tree.delete(i)

    # Draw the cities
    for i, (cx, cy, cradius) in enumerate(city_centers):
        draw_city(cx, cy, cradius, i)
    # Draw the base stations
    for station in base_stations:
        draw_base_station(station)

def draw_city(cx, cy, cradius, city_index):
    canvas.create_text(cx, cy - 18, text=chr(65 + city_index), font=("Arial", 14, "bold"), fill='red', tags=("city","zoomable"))
    canvas.create_image(cx, cy, image=city_icon_tk, tags=("city","zoomable"))
    canvas.create_oval(cx - cradius, cy - cradius, cx + cradius, cy + cradius, outline='red', width=3, tags=("city","zoomable"))
    city_tree.insert("", tk.END, values=(chr(65 + city_index), round(cx), round(cy), round(cradius * METERS_PER_PIXEL)))

def draw_base_station(station):
    idx, x, y, radius, position = station.values()
    color = 'black' if position == 'IN' else 'blue'
    canvas.create_oval(x - 1, y - 1, x + 1, y + 1, fill=color, tags=("base_station", "zoomable"))
    canvas.create_oval(x - radius, y - radius, x + radius, y + radius, outline=color, width=2, tags=("base_station", "zoomable"))
    canvas.create_text(x, y - 10, text=str(idx), font=("Arial", 10), fill='black', tags=("base_station", "zoomable"))
    tree.insert("", tk.END, values=(idx, round(x), round(y), round(radius * METERS_PER_PIXEL), position))

def draw_random_points():
    #Draws random base stations and cities on the canvas based on user parameters.
    canvas.delete("base_station")
    canvas.delete("city")
    canvas.delete("triangulation")
    triangulation_table.delete(*triangulation_table.get_children())
    clear_neighbors()
    clear_highlight()
    base_stations.clear()

    global current_zoom
    if current_zoom != 1.0:
        reset_scale = 1.0 / current_zoom  # Calculate the inverse of the current zoom to reset it
        canvas.scale("zoomable", 0, 0, reset_scale, reset_scale)
        current_zoom = 1.0

    global METERS_PER_PIXEL
    global PIXELS_PER_METER
    PIXELS_PER_METER = get_scale_value()
    METERS_PER_PIXEL = 1 / PIXELS_PER_METER

    # Convert input parameters from meters to pixels
    min_city_radius_pixels = round(min_city_radius_var.get() * PIXELS_PER_METER*1000)
    max_city_radius_pixels = round(max_city_radius_var.get() * PIXELS_PER_METER*1000)

    # Gathering and preparing city parameters
    min_cities = int(min_cities_var.get())
    max_cities = int(max_cities_var.get())
    if min_cities > max_cities:
        tk.messagebox.showerror("Error", "Minimum number of cities cannot be greater than the maximum number.")
        return
    num_cities = random.randint(min_cities, max_cities)
    min_city_radius = int(min_city_radius_var.get())
    max_city_radius = int(max_city_radius_var.get())
    if min_city_radius > max_city_radius:
        tk.messagebox.showerror("Error", "Minimum city radius cannot be greater than the maximum city radius.")
        return

    # Checking if cities should be retained or drawn a new
    if not keep_cities_var.get() or not city_centers:
        city_centers.clear()  # Also delete existing cities if they were drawn previously

        # Drawing the cities
        for _ in range(num_cities):
            tries = 0
            city_radius = random.randint(min_city_radius_pixels, max_city_radius_pixels)
            # Finding a position for the city such that it doesn't overlap with other cities
            while tries < 1000:
                x = random.randint(10 + city_radius, 10 + SQUARE_SIZE - city_radius)
                y = random.randint(10 + city_radius, 10 + SQUARE_SIZE - city_radius)
                if not are_points_within_range(x, y, city_centers, 2 * city_radius):
                    city_centers.append((x, y, city_radius))
                    break
                tries += 1
            if tries == 1000:
                tk.messagebox.showerror("Error", "Failed to place all the stations/cities. Try adjusting parameters.")
                return

    existing_points = []    # List to store the locations of base stations to prevent overlap
    num_points = int(num_points_var.get())
    percentage_in_city = int(percentage_in_city_var.get())
    percentage_outside = int(percentage_outside_var.get())
    if percentage_in_city + percentage_outside != 100:
        tk.messagebox.showerror("Error", "The sum of percentages should be 100%.")
        return

    # Calculate the number of base stations inside and outside of cities
    stations_in_city = int((percentage_in_city / 100) * num_points)
    stations_outside = num_points - stations_in_city
    stations_per_city = stations_in_city // num_cities

    # Place base stations inside the cities
    for cx, cy, cradius in city_centers:
        for _ in range(stations_per_city):
            circle_radius_in_city_pixels = random.randint(
                round(min_radius_in_city_var.get() * PIXELS_PER_METER*100),
                round(max_radius_in_city_var.get() * PIXELS_PER_METER*100)
            )
            tries = 0
            while tries < 1000:
                angle = 2 * math.pi * random.random()   # Randomly choose an angle and distance to position the base station within the city
                distance_from_center = cradius * random.random()
                x = cx + distance_from_center * math.cos(angle)
                y = cy + distance_from_center * math.sin(angle)
                # Ensure this base station doesn't overlap with other base stations
                if not are_points_within_range(x, y, existing_points, 0, inside_multiplier_var.get() * circle_radius_in_city_pixels):
                    existing_points.append((x, y))
                    base_stations.append({"id": len(base_stations), "x": x, "y": y, "radius": circle_radius_in_city_pixels, "position": "IN"})
                    break
                tries += 1

    # Place base stations outside the cities
    for _ in range(stations_outside):
        circle_radius_outside_pixels = random.randint(
            round(min_radius_outside_var.get() * PIXELS_PER_METER*1000),
            round(max_radius_outside_var.get() * PIXELS_PER_METER*1000)
        )
        tries = 0
        while tries < 1000:
            # Randomly choose a position for the base station
            x = random.randint(10 + circle_radius_outside_pixels, 10 + SQUARE_SIZE - circle_radius_outside_pixels)
            y = random.randint(10 + circle_radius_outside_pixels, 10 + SQUARE_SIZE - circle_radius_outside_pixels)
            # Ensure this base station doesn't overlap with cities or other base stations
            if not any(are_points_within_range(x, y, [(cx, cy)], cr) for cx, cy, cr in city_centers) and not are_points_within_range(x, y, existing_points, 0, outside_multiplier_var.get() * circle_radius_outside_pixels):
                existing_points.append((x, y))
                base_stations.append({"id": len(base_stations), "x": x, "y": y, "radius": circle_radius_outside_pixels, "position": "OUT"})
                break
            tries += 1

    # Draw cities and base stations
    for i, (cx, cy, cradius) in enumerate(city_centers):
        draw_city(cx, cy, cradius, i)
    for station in base_stations:
        draw_base_station(station)

    # Clearing previous records in the table
    for i in tree.get_children():
            tree.delete(i)
    for item in selected_tree.get_children():
        selected_tree.delete(item)

    for station in base_stations:
        x = station["x"]
        y = station["y"]
        idx = station["id"]
        radius = round(station["radius"] * METERS_PER_PIXEL)
        position = station["position"]
        tree.insert("", tk.END, values=(idx, round(x), round(y), radius, position))

    for i in city_tree.get_children():
        city_tree.delete(i)
    for i, (cx, cy, cradius) in enumerate(city_centers):
        city_tree.insert("", tk.END, values=(chr(65 + i), round(cx), round(cy), round(cradius*METERS_PER_PIXEL)))

def perform_delaunay_triangulation():
    canvas.delete("triangulation")
    triangulation_table.delete(*triangulation_table.get_children())

    points = [(station["x"], station["y"]) for station in base_stations]
    tri = Delaunay(points)

    global triangulated_neighbors
    triangulated_neighbors = {i: set() for i in range(len(base_stations))}

    for simplex in tri.simplices:
        for i in range(3):
            start_index, end_index = simplex[i], simplex[(i + 1) % 3]
            canvas.create_line(points[start_index], points[end_index], fill='black', tags=("triangulation", "zoomable"))

            triangulated_neighbors[start_index].add(end_index)
            triangulated_neighbors[end_index].add(start_index)

    for station_id, neighbors in triangulated_neighbors.items():
        neighbors_formatted = ', '.join(str(n) for n in sorted(neighbors))
        triangulation_table.insert("", 'end', values=(station_id, neighbors_formatted))

def clear_triangulation():
    canvas.delete("triangulation")

def create_excel_file(triangulated_neighbors, real_neighbors):
    filename = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        title="Save as..."
    )
    if not filename:
        return

    wb = Workbook()
    wb.remove(wb.active)

    ws_triangulated = wb.create_sheet("Triangulated Neighbors")
    ws_triangulated.append(['Station ID', 'Neighbors'])
    for station_id, neighbors in triangulated_neighbors.items():
        neighbors_formatted = ', '.join(str(n) for n in neighbors)
        ws_triangulated.append([station_id, neighbors_formatted])

    ws_real = wb.create_sheet("Real Neighbors")
    ws_real.append(['Station ID', 'Neighbors'])
    for station_id, neighbors in real_neighbors.items():
        neighbors_formatted = ', '.join(str(n) for n in neighbors)
        ws_real.append([station_id, neighbors_formatted])

    wb.save(filename)
    tk.messagebox.showinfo('Export to Excel', f'Data exported successfully to {filename}')


def export_neighbors_to_excel():
    create_excel_file(triangulated_neighbors, real_neighbors)


def draw_real_connection(canvas, station1, station2):
    """Draws a line of connection between two stations."""
    canvas.create_line(station1['x'], station1['y'], station2['x'], station2['y'], fill='#36d10c',width=2, tags=("neighbors", "zoomable"))

def is_clear_path_between_stations(station1, station2, other_stations):
    """
    Checks if there is a clear path between two stations, ensuring no other station's radius
    intersects the line connecting these two stations.

    :param station1: Dictionary with data of the first station (coordinates and radius).
    :param station2: Dictionary with data of the second station (coordinates and radius).
    :param other_stations: List of dictionaries with data of other stations.
    :return: True if the path between stations is clear, False otherwise.
    """
    x1, y1 = station1['x'], station1['y']  # Coordinates of the first station
    x2, y2 = station2['x'], station2['y']  # Coordinates of the second station

    for station in other_stations:
        if station == station1 or station == station2:
            continue  # Skip the stations we are checking

        # Coordinates and radius of the other station
        sx, sy, sradius = station['x'], station['y'], station['radius']

        # Check if the line intersects with the station's radius
        if is_line_intersecting_circle((x1, y1), (x2, y2), (sx, sy), sradius):
            return False  # Path is blocked

    return True  # Path is clear

def is_line_intersecting_circle(p1, p2, center, radius):
    """
    Determines if a line segment between two points intersects a circle of given radius.

    :param p1: Coordinates of the first point of the line segment.
    :param p2: Coordinates of the second point of the line segment.
    :param center: Coordinates of the center of the circle.
    :param radius: Radius of the circle.
    :return: True if the line segment intersects the circle, False otherwise.
    """
    cx, cy = center  # Center of the circle
    ax, ay = p1      # First point of the line segment
    bx, by = p2      # Second point of the line segment

    # Vector from p1 to p2
    ab = (bx - ax, by - ay)
    # Vector from p1 to circle center
    ac = (cx - ax, cy - ay)
    # Dot product of ab and ac
    ab_dot_ac = ab[0] * ac[0] + ab[1] * ac[1]

    # Finding the projection of the center onto the line (parameter t of the line equation)
    t = ab_dot_ac / (ab[0] ** 2 + ab[1] ** 2)

    # Finding the closest point on the line to the circle center
    closest_x = ax + ab[0] * t
    closest_y = ay + ab[1] * t

    # Check if the closest point is within the line segment
    if not (min(ax, bx) <= closest_x <= max(ax, bx) and min(ay, by) <= closest_y <= max(ay, by)):
        return False

    # Check if the distance from the closest point to the circle center is less than the radius
    return math.hypot(closest_x - cx, closest_y - cy) < radius



def update_table_with_real_neighbors(neighbors_table, station_id, neighbors):
    """Updates the table with information on the real neighbors of a station."""
    neighbors_formatted = ', '.join(str(n) for n in sorted(neighbors))
    neighbors_table.insert("", 'end', values=(station_id, neighbors_formatted))

def find_real_neighbors(base_stations, station_neighbors, alpha, beta, canvas, neighbors_table):
    global real_neighbors
    real_neighbors = {station['id']: set() for station in base_stations}

    for station in base_stations:
        x1, y1, radius1, position1 = station['x'], station['y'], station['radius'], station['position']
        for neighbor_id in station_neighbors[station['id']]:
            neighbor = next((s for s in base_stations if s['id'] == neighbor_id), None)
            if neighbor:
                x2, y2, radius2, position2 = neighbor['x'], neighbor['y'], neighbor['radius'], neighbor['position']

                # Calculate the distance between the two stations
                distance = math.hypot(x2 - x1, y2 - y1)

                # Determine the coefficient to use based on the positions of the stations
                coefficient = alpha if position1 == 'IN' and position2 == 'IN' else beta

                # Check if the stations satisfy the condition based on coefficient and radius sum
                # and if no other station is in between them
                if (distance <= coefficient * (radius1 + radius2) and is_clear_path_between_stations(station, neighbor, base_stations)):
                    real_neighbors[station['id']].add(neighbor_id)
                    draw_real_connection(canvas, station, neighbor)

    # Update the table with the real neighbors
    for station_id, neighbors in real_neighbors.items():
        update_table_with_real_neighbors(neighbors_table, station_id, neighbors)

    return real_neighbors


def on_find_neighbors_button_click():
    """Handles the click event of the button to find real neighbors."""
    try:
        alpha = alpha_var.get()  # Adjust the alpha value as needed
        beta = beta_var.get()  # Adjust the beta value as needed
        real_neighbors = find_real_neighbors(base_stations, triangulated_neighbors, alpha, beta, canvas, neighbors_table)
        messagebox.showinfo("Information", "Real neighbors have been successfully found and displayed.")
    except Exception as e:
        messagebox.showerror("Error", str(e))




def clear_neighbors():
    canvas.delete("neighbors")
    neighbors_table.delete(*neighbors_table.get_children())

alpha_var = tk.DoubleVar(value=1.2)
beta_var = tk.DoubleVar(value=1.2)

num_points_var = tk.IntVar(value=100)
min_radius_in_city_var = tk.IntVar(value=20)
max_radius_in_city_var = tk.IntVar(value=30)
min_radius_outside_var = tk.IntVar(value=2)
max_radius_outside_var = tk.IntVar(value=7)
min_cities_var = tk.IntVar(value=2)
max_cities_var = tk.IntVar(value=5)
inside_multiplier_var = tk.DoubleVar(value=1.4)
min_city_radius_var = tk.IntVar(value=7)
max_city_radius_var = tk.IntVar(value=15)
outside_multiplier_var = tk.DoubleVar(value=1.4)
percentage_in_city_var = tk.IntVar(value=90)
percentage_outside_var = tk.IntVar(value=10)
keep_cities_var = tk.BooleanVar()

scale_frame = ctk.CTkFrame(root, border_width=2, corner_radius=7,border_color="black")
scale_var = tk.StringVar(value="Current Scale: 100000m | Select:")
scale_label = ctk.CTkLabel(scale_frame, textvariable=scale_var, font=("Arial", 16))
scale_label.grid(row=0, column=0, padx=5, pady=5, sticky="ew")
scale_selection_var = tk.StringVar(value='100000m')
scale_menu = ttk.Combobox(scale_frame, textvariable=scale_selection_var, values=list(SCALE_OPTIONS.keys()), state="readonly", font=("Arial", 16), width=10)
scale_menu.grid(row=0, column=1, padx=15, pady=5, sticky="ew")

def get_scale_value():
    selected_scale = scale_selection_var.get()
    return SCALE_OPTIONS[selected_scale]
def on_scale_select(event):
    selected_scale = scale_selection_var.get()
    scale_var.set(f"Current Scale: {selected_scale} | Select:")
    scale_label.focus()
scale_menu.bind('<<ComboboxSelected>>', on_scale_select)
scale_frame_window = canvas.create_window(screen_width - 310, screen_height - 160, window=scale_frame, anchor='nw')

def update_scale_frame_position(event):
    canvas_width = canvas.winfo_width()
    canvas_height = canvas.winfo_height()
    canvas.coords(scale_frame_window, canvas_width - scale_frame.winfo_reqwidth() - 10, canvas_height - scale_frame.winfo_reqheight() - 10)
root.bind('<Configure>', update_scale_frame_position)


configuration_frame = ctk.CTkFrame(root, border_width=3)
configuration_frame.grid(row=0, column=1, padx=5, pady=7, sticky="nw")
ctk.CTkLabel(configuration_frame, text="CONFIGURATION FOR GENERATION:",fg_color="grey",corner_radius=5, font=ctk.CTkFont(size=12, weight="bold")).grid(row=0, columnspan=3, sticky="ew")

cities_frame = ctk.CTkFrame(configuration_frame, border_width=3)
cities_frame.grid(row=1, column=1, padx=5, pady=5, sticky="new")

ctk.CTkLabel(cities_frame, text="City Count Range",fg_color="grey",corner_radius=5, text_color="black",font=ctk.CTkFont(size=12, weight="bold")).grid(row=0, column=0,padx=5, pady=5,sticky="nsew")
cities_count_frame = ctk.CTkFrame(cities_frame,border_width=2)
cities_count_frame.grid(row=1, column=0, padx=15, pady=5, sticky="nsew")
ctk.CTkLabel(cities_count_frame, text="MIN COUNT:").grid(row=0, column=0, sticky="e", padx=5, pady=5)
ctk.CTkEntry(cities_count_frame, textvariable=min_cities_var, width=35,corner_radius=5).grid(row=0, column=1, padx=5, pady=5)
ctk.CTkLabel(cities_count_frame, text="MAX COUNT:").grid(row=1, column=0, sticky="e", padx=5, pady=5)
ctk.CTkEntry(cities_count_frame, textvariable=max_cities_var, width=35,corner_radius=5).grid(row=1, column=1, padx=5, pady=5)

ctk.CTkLabel(cities_frame, text="City Radius Range",fg_color="grey",corner_radius=5, text_color="black",font=ctk.CTkFont(size=12, weight="bold")).grid(row=2, column=0,padx=5, pady=5,sticky="nsew")
city_radius_frame = ctk.CTkFrame(cities_frame, border_width=2)
city_radius_frame.grid(row=3, column=0, padx=15, pady=5, sticky="nsew")
ctk.CTkLabel(city_radius_frame, text="MIN [ Km ]:").grid(row=0, column=0, sticky="e", padx=5, pady=5)
ctk.CTkEntry(city_radius_frame, textvariable=min_city_radius_var, width=35,corner_radius=5).grid(row=0, column=1, sticky="w", padx=5, pady=5)
ctk.CTkLabel(city_radius_frame, text="MAX [ Km ]:").grid(row=1, column=0, sticky="e", padx=5, pady=5)
ctk.CTkEntry(city_radius_frame, textvariable=max_city_radius_var, width=35,corner_radius=5).grid(row=1, column=1, sticky="w", padx=5, pady=5)

ctk.CTkLabel(cities_frame, text="City Information",fg_color="grey",corner_radius=5, text_color="black",font=ctk.CTkFont(size=12, weight="bold")).grid(row=4, columnspan=1, sticky="nsew")
city_info_frame = ctk.CTkFrame(cities_frame)
city_info_frame.grid(row=5, column=0, padx=5, pady=5, sticky="ewn")
city_tree = ttk.Treeview(city_info_frame, height=2)
city_tree["columns"] = ("ID", "X", "Y", "Radius")
city_tree.column("#0", width=0, stretch=tk.NO)
city_tree.column("ID", anchor=tk.W, width=40)
city_tree.column("X", anchor=tk.W, width=40)
city_tree.column("Y", anchor=tk.W, width=40)
city_tree.column("Radius", anchor=tk.W, width=55)
city_tree.heading("#0", text="", anchor=tk.W)
city_tree.heading("ID", text="ID", anchor=tk.W)
city_tree.heading("X", text="X", anchor=tk.W)
city_tree.heading("Y", text="Y", anchor=tk.W)
city_tree.heading("Radius", text="Radius", anchor=tk.W)
city_tree.grid(row=0, column=0, columnspan=3, padx=5, pady=5, sticky='new')


bts_info_frame = ctk.CTkFrame(configuration_frame, border_width=3)
bts_info_frame.grid(row=1, column=2, padx=5, pady=7, sticky="news")

ctk.CTkLabel(bts_info_frame, text="Info about generated BTS:",fg_color="grey",corner_radius=5, font=ctk.CTkFont(size=12, weight="bold")).grid(row=0, columnspan=3, sticky="ew")

bts_table_frame = ctk.CTkFrame(bts_info_frame,border_width=2)
bts_table_frame.grid(row=1, column=0,  padx=5, pady=5, sticky="nw")

tree = ttk.Treeview(bts_table_frame, height=5)
tree["columns"] = ("ID", "X", "Y", "Radius", "Position")
tree.column("#0", width=0, stretch=tk.NO)
tree.column("ID", anchor=tk.W, width=30)
tree.column("X", anchor=tk.W, width=30)
tree.column("Y", anchor=tk.W, width=30)
tree.column("Radius", anchor=tk.W, width=45)
tree.column("Position", anchor=tk.W, width=50)
tree.heading("#0", text="", anchor=tk.W)
tree.heading("ID", text="ID", anchor=tk.W)
tree.heading("X", text="X", anchor=tk.W)
tree.heading("Y", text="Y", anchor=tk.W)
tree.heading("Radius", text="Radius", anchor=tk.W)
tree.heading("Position", text="Position", anchor=tk.W)
tree.grid(row=1, column=0, columnspan=3, padx=10, pady=13, sticky='ew')

bts_info_scrollbar = tk.Scrollbar(bts_table_frame, orient="vertical", command=tree.yview)
bts_info_scrollbar.grid(row=1, column=4, sticky='ns')
tree.configure(yscrollcommand=bts_info_scrollbar.set)

stations_frame = ctk.CTkFrame(configuration_frame, border_width=3)
stations_frame.grid(row=1, column=0, padx=5, pady=5, sticky="nsew")

stations_count_frame = ctk.CTkFrame(stations_frame)
stations_count_frame.grid(row=0, columnspan=2, padx=5, pady=5, sticky="news")

ctk.CTkLabel(stations_count_frame, text="STATION COUNT:", font=ctk.CTkFont(size=15, weight="bold")).grid(row=0, column=0, padx=5, pady=5,sticky="ew")
ctk.CTkEntry(stations_count_frame, textvariable=num_points_var).grid(row=0, column=1, padx=5, pady=5,sticky="ew")

in_city_frame = ctk.CTkFrame(stations_frame, border_width=2)
in_city_frame.grid(row=1, column=0, columnspan=3, padx=10, pady=5, sticky="ew")
ctk.CTkLabel(in_city_frame, text="Radius range for IN-city stations:",fg_color="grey",corner_radius=5, font=ctk.CTkFont(size=12, weight="bold")).grid(row=0, columnspan=3, sticky="ew")
radius_in_city_frame = ctk.CTkFrame(in_city_frame)
radius_in_city_frame.grid(row=1, columnspan=3, padx=5, pady=5, sticky="ew")

ctk.CTkLabel(radius_in_city_frame, text="MIN:\n[ *100 m ]").grid(row=0, column=0, sticky="w", padx=5, pady=5)
radius_in_city_scale = RangeSliderH(radius_in_city_frame, [min_radius_in_city_var, max_radius_in_city_var],padX=11, min_val=1, max_val=30, step_size=1, Width=200, Height=60, bar_radius=7, bgColor="#CFCFCF")
radius_in_city_scale.grid(row=0, column=1, pady=5, sticky="ew")
ctk.CTkLabel(radius_in_city_frame, text="MAX:\n[ *100m ]").grid(row=0, column=2, sticky="e", padx=5, pady=5)

in_city_multiplier_frame = ctk.CTkFrame(in_city_frame, border_width=2)
in_city_multiplier_frame.grid(row=4, columnspan=3, padx=5,pady=5)

ctk.CTkLabel(in_city_frame, text="Station Radius Overlapping:", fg_color="grey",corner_radius=5,font=ctk.CTkFont(size=12, weight="bold")).grid(row=3,columnspan=3, sticky="ew")
ctk.CTkLabel(in_city_multiplier_frame, text="MIN").grid(row=0, column=0, sticky="w", padx=5, pady=5)
outside_multiplier_slider = ctk.CTkSlider(in_city_multiplier_frame, from_=2, to=1, variable=inside_multiplier_var, number_of_steps=20, width=150, height=10, border_width=3)
outside_multiplier_slider.grid(row=0, column=1, pady=5, sticky="ew")
ctk.CTkLabel(in_city_multiplier_frame, text="MAX").grid(row=0, column=2, sticky="e", padx=5, pady=5)

outside_city_frame = ctk.CTkFrame(stations_frame, border_width=2)
outside_city_frame.grid(row=2, column=0, columnspan=3, padx=10, pady=5, sticky="nsew")

ctk.CTkLabel(outside_city_frame, text="Radius range for OUT-of-city stations:",fg_color="grey",corner_radius=5, font=ctk.CTkFont(size=12, weight="bold")).grid(row=0,columnspan=3, sticky="ew")
radius_frame = ctk.CTkFrame(outside_city_frame)
radius_frame.grid(row=1, columnspan=3, padx=5, pady=5, sticky="ew")
ctk.CTkLabel(radius_frame, text="MIN:\n[ Km ]").grid(row=0, column=0, sticky="w", padx=5, pady=5)
radius_outside_city_scale = RangeSliderH(radius_frame, [min_radius_outside_var, max_radius_outside_var], padX=11, min_val=1, max_val=20, step_size=1, Width=250, Height=60, bar_radius=7, bgColor="#CFCFCF")
radius_outside_city_scale.grid(row=0, column=1, pady=5, sticky="ew")
ctk.CTkLabel(radius_frame, text="MAX:\n[ Km ]").grid(row=0, column=2, sticky="e", padx=5, pady=5)

multiplier_frame = ctk.CTkFrame(outside_city_frame, border_width=2)
multiplier_frame.grid(row=4, columnspan=3, padx=5,pady=5)

ctk.CTkLabel(outside_city_frame, text="Station Radius Overlapping:", fg_color="grey",corner_radius=5,font=ctk.CTkFont(size=12, weight="bold")).grid(row=3,columnspan=3, sticky="ew")
ctk.CTkLabel(multiplier_frame, text="MIN").grid(row=0, column=0, sticky="w", padx=5, pady=5)
outside_multiplier_slider = ctk.CTkSlider(multiplier_frame, from_=2, to=1, variable=outside_multiplier_var, number_of_steps=20, width=150, height=10, border_width=3)
outside_multiplier_slider.grid(row=0, column=1, pady=5, sticky="ew")
ctk.CTkLabel(multiplier_frame, text="MAX").grid(row=0, column=2, sticky="e", padx=5, pady=5)

stations_percent_frame = ctk.CTkFrame(stations_frame, border_width=3)
stations_percent_frame.grid(row=5, column=0, padx=10, pady=5, sticky="ew")
inside_percentage_string = tk.StringVar(value="INSIDE CITY [ 90 % ]")
outside_percentage_string = tk.StringVar(value="OUTSIDE CITY [ 10 % ]")
ctk.CTkLabel(stations_percent_frame, text="Location percentages for BTS generation:",fg_color="grey",corner_radius=5, font=ctk.CTkFont(size=12, weight="bold")).grid(row=0,columnspan=2, sticky="news",padx=5, pady=5)
ctk.CTkLabel(stations_percent_frame, textvariable=inside_percentage_string, font=ctk.CTkFont(size=12, weight="bold")).grid(row=1, column=0, sticky="e", padx=5, pady=5)
inside_city_slider = ctk.CTkSlider(stations_percent_frame, from_=0, to=100, variable=percentage_in_city_var, number_of_steps=100, width=150, height=10, border_width=3)
inside_city_slider.grid(row=1, column=1, padx=5, pady=5, sticky="ew")
ctk.CTkLabel(stations_percent_frame, textvariable=outside_percentage_string, font=ctk.CTkFont(size=12, weight="bold")).grid(row=2, column=0, sticky="e", padx=5, pady=5)
outside_city_slider = ctk.CTkSlider(stations_percent_frame, from_=0, to=100, variable=percentage_outside_var, number_of_steps=100, width=150, height=10, border_width=3)
outside_city_slider.grid(row=2, column=1, padx=5, pady=5, sticky="ew")

def update_outside_slider(*args):
    outside_val = 100 - percentage_in_city_var.get()
    percentage_outside_var.set(outside_val)
    outside_percentage_string.set(f"Outside sity [ {outside_val} % ]")
def update_inside_slider(*args):
    inside_val = 100 - percentage_outside_var.get()
    percentage_in_city_var.set(inside_val)
    inside_percentage_string.set(f"Inside sity [ {inside_val} % ]")

percentage_in_city_var.trace("w", update_outside_slider)
percentage_outside_var.trace("w", update_inside_slider)

selected_bts_info_frame = ctk.CTkFrame(bts_info_frame, border_width=3)
selected_bts_info_frame.grid(row=5, column=0, padx=5, pady=5, sticky="news")


controls_frame = ctk.CTkFrame(selected_bts_info_frame)
controls_frame.grid(row=0, column=0, columnspan=3, padx=5, pady=5, sticky="news")

station_id_entry_var = tk.StringVar()
ctk.CTkLabel(controls_frame, text="Station ID:").grid(row=0, column=0, sticky="w", padx=5, pady=5)
station_id_entry = ctk.CTkEntry(controls_frame, textvariable=station_id_entry_var, width=50)
station_id_entry.grid(row=0, column=1, padx=5, pady=5)

find_button = ctk.CTkButton(controls_frame, text="Find", command=find_station, width=30)
find_button.grid(row=0, column=2, padx=5, pady=5)

delete_button = ctk.CTkButton(selected_bts_info_frame, text="Delete BTS", command=delete_selected_station)
delete_button.grid(row=2, column=1, padx=5, pady=5, sticky="news")

clear_highlight_btn = ctk.CTkButton(selected_bts_info_frame, text="Clear all Selected BTS", command=clear_highlight)
clear_highlight_btn.grid(row=3, column=1, pady=10, padx=5,sticky="news")

selected_tree = ttk.Treeview(selected_bts_info_frame, height=3)
selected_tree["columns"] = ("ID", "X", "Y", "Radius", "Position")
selected_tree.column("#0", width=0, stretch=tk.NO)
selected_tree.column("ID", anchor=tk.W, width=30)
selected_tree.column("X", anchor=tk.W, width=30)
selected_tree.column("Y", anchor=tk.W, width=30)
selected_tree.column("Radius", anchor=tk.W, width=45)
selected_tree.column("Position", anchor=tk.W, width=50)
selected_tree.heading("#0", text="", anchor=tk.W)
selected_tree.heading("ID", text="ID", anchor=tk.W)
selected_tree.heading("X", text="X", anchor=tk.W)
selected_tree.heading("Y", text="Y", anchor=tk.W)
selected_tree.heading("Radius", text="Radius", anchor=tk.W)
selected_tree.heading("Position", text="Position", anchor=tk.W)
selected_tree.grid(row=1, column=0, columnspan=3, padx=10, pady=10, sticky='ew')

controls_frame = ctk.CTkFrame(configuration_frame,border_width=3)
controls_frame.grid(row=1, column=1, pady=5, padx=5, sticky="ews")
ctk.CTkCheckBox(controls_frame, text="Keep cities on map", variable=keep_cities_var).grid(row=0, column=0, pady=5, padx=10,sticky="news")
ctk.CTkButton(controls_frame, text="\n          Apply          \n", command=draw_random_points).grid(row=1, column=0,pady=5, padx=10,sticky="news")

file_frame = ctk.CTkFrame(bts_info_frame)
file_frame.grid(row=3, column=0, padx=5, pady=5)
save_button = ctk.CTkButton(file_frame, text="\n   Save Configuration   \n", command=save_configuration_as)
save_button.grid(row=0, column=0, pady=5, padx=5,sticky="news")
load_button = ctk.CTkButton(file_frame, text="\n   Load Configuration   \n", command=load_configuration)
load_button.grid(row=1, column=0, pady=5, padx=5,sticky="news")

processing_frame = ctk.CTkFrame(root, border_width=3)
processing_frame.grid(row=1, column=1, padx=5, pady=5, sticky="news")

ctk.CTkLabel(processing_frame, text="CONFIGURATION PROCESSING:",fg_color="grey",corner_radius=5, font=ctk.CTkFont(size=12, weight="bold")).grid(row=0, columnspan=4, sticky="ew")

delaunay_triangulation_frame = ctk.CTkFrame(processing_frame, border_width=3)
delaunay_triangulation_frame.grid(row=1, column=0, padx=5, pady=5, sticky="news")

processing_control_frame = ctk.CTkFrame(delaunay_triangulation_frame,border_width=2)
processing_control_frame.grid(row=3, column=0,  padx=5, pady=5, sticky="nw")
triangulate_button = ctk.CTkButton(processing_control_frame, text="Triangulate", command=perform_delaunay_triangulation, width=80, height=45)
triangulate_button.grid(row=0, column=0, padx=5, pady=5)
clear_triangulation_button = ctk.CTkButton(processing_control_frame, text="Clear", command=clear_triangulation,width=80, height=45)
clear_triangulation_button.grid(row=1, column=0, padx=5, pady=5)
export_excel_button = ctk.CTkButton(processing_control_frame, text="Export", command=export_neighbors_to_excel,width=80, height=45)
export_excel_button.grid(row=2, column=0, padx=5, pady=5, sticky="ew")


triangulation_frame = ctk.CTkFrame(processing_frame, border_width=3)
triangulation_frame.grid(row=1, column=1, padx=5, pady=5, sticky="news")
ctk.CTkLabel(triangulation_frame, text="Delaunay Treangulation:",fg_color="grey",corner_radius=5, font=ctk.CTkFont(size=12, weight="bold")).grid(row=0, column=0, sticky="ew")
processing_table_frame = ctk.CTkFrame(triangulation_frame,border_width=2)
processing_table_frame.grid(row=1, column=0,  padx=5, pady=5, sticky="nw")

triangulation_table = ttk.Treeview(processing_table_frame,height=7)
triangulation_table['columns'] = ('Station ID', 'Neighbors')
triangulation_table.column("#0", width=0, stretch=tk.NO)
triangulation_table.column("Station ID", anchor=tk.CENTER, width=60)
triangulation_table.column("Neighbors", anchor=tk.CENTER, width=200)
triangulation_table.heading("#0", text='', anchor=tk.CENTER)
triangulation_table.heading("Station ID", text="Station ID", anchor=tk.CENTER)
triangulation_table.heading("Neighbors", text="Neighbors", anchor=tk.CENTER)
triangulation_table.grid(row=1, column=1, padx=5, pady=5)


delaunay_triangulation_frame = ctk.CTkFrame(processing_frame, border_width=3)
delaunay_triangulation_frame.grid(row=1, column=2, padx=5, pady=5, sticky="news")

processing_control_frame = ctk.CTkFrame(delaunay_triangulation_frame,border_width=2)
processing_control_frame.grid(row=0, column=0,  padx=5, pady=5, sticky="nwe")
ctk.CTkLabel(processing_control_frame, text="   [ IN ] coef :").grid(row=0, column=0, sticky="ew", padx=5, pady=5)
ctk.CTkEntry(processing_control_frame, textvariable=alpha_var, width=35,corner_radius=5).grid(row=0, column=1, sticky="e", padx=5, pady=5)
ctk.CTkLabel(processing_control_frame, text="[ OUT ] coef :").grid(row=1, column=0, sticky="ew", padx=5, pady=5)
ctk.CTkEntry(processing_control_frame, textvariable=beta_var, width=35,corner_radius=5).grid(row=1, column=1, sticky="e", padx=5, pady=5)

triangulate_button = ctk.CTkButton(delaunay_triangulation_frame, text="Find", command=on_find_neighbors_button_click, width=110, height=35)
triangulate_button.grid(row=2, column=0, padx=5, pady=5)
triangulate_button = ctk.CTkButton(delaunay_triangulation_frame, text="Clear", command=clear_neighbors, width=110, height=35)
triangulate_button.grid(row=3, column=0, padx=5, pady=5)


triangulation_frame = ctk.CTkFrame(processing_frame, border_width=3)
triangulation_frame.grid(row=1, column=3, padx=5, pady=5, sticky="news")
ctk.CTkLabel(triangulation_frame, text="Station Neighbors:",fg_color="grey",corner_radius=5, font=ctk.CTkFont(size=12, weight="bold")).grid(row=0, column=3, sticky="ew")
processing_table_frame = ctk.CTkFrame(triangulation_frame,border_width=2)
processing_table_frame.grid(row=1, column=3,  padx=5, pady=5, sticky="nw")

neighbors_table = ttk.Treeview(processing_table_frame,height=7)
neighbors_table['columns'] = ('Station ID', 'Neighbors')
neighbors_table.column("#0", width=0, stretch=tk.NO)
neighbors_table.column("Station ID", anchor=tk.CENTER, width=60)
neighbors_table.column("Neighbors", anchor=tk.CENTER, width=200)
neighbors_table.heading("#0", text='', anchor=tk.CENTER)
neighbors_table.heading("Station ID", text="Station ID", anchor=tk.CENTER)
neighbors_table.heading("Neighbors", text="Neighbors", anchor=tk.CENTER)
neighbors_table.grid(row=1, column=1, padx=5, pady=5)

draw_random_points()
root.mainloop()